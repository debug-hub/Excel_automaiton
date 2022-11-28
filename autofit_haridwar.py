import openpyxl
from openpyxl import *
import re


path = r"D:\SequelString\[pcr][HCL]\PIR_Autofit UTR_100489_01042021.xlsx"
excel = load_workbook(path)
excel_1 = load_workbook(path,data_only=True)

############################

path2 = r"D:\SequelString\[pcr][HCL]\AUTOFIT_MASTER.xlsx"
wb = Workbook()
MASTER= wb['Sheet']
MASTER.title= "Master BOM Hier"
MASTER['A1']="BOM HIERARCHY"
MASTER['B1']="MASTER"
MASTER['C1']="DIRECT SUP"
MASTER['D1']="PLANT"
MASTER['E1']="FREQUENCY" 
MASTER['F1']="FROM DATE"
MASTER['G1']="TO DATE"
MASTER['H1']="PURCHASE Group"
MASTER['I1']="VALUE"
MASTER['J1']="PERCENATGE"


############################

path3 = "D:\SequelString\[pcr][HCL]\AUTOFIT_Hierarchy.xlsx"
wb2 = Workbook()
Hierarchy = wb2['Sheet']

Hierarchy['A1']="BOM HIERARCHY"
Hierarchy['B1']="DIRECT SUP"
Hierarchy['C1']="PURCHASE Group"
Hierarchy['D1']="PLANT"
Hierarchy['E1']="FROM DATE"
Hierarchy['F1']="TO DATE"
Hierarchy['G1']="partcode"
Hierarchy['H1']="partcode type1(OE)"
Hierarchy['I1']="partcode Level2"
Hierarchy['J1']="partcode type2(RM/BOP/VTV/INM)"
Hierarchy['K1']="partcode Level3"
Hierarchy['L1']="partcode type3(RM/BOP/VTV/INM)"



############################


sheet_1 = excel_1['PO']


sheet2_1 = excel_1['TOP SHEET']


sheet3_1 = excel_1['BOM']


sheet4_1 = excel_1['VTV Price']


sheet5_1 = excel_1['Conversion']

############################


def master(MASTER,wb,path2,part_code, vendor,plant,xx,CONV_COST):
    rows = MASTER.max_row

        
    c0 = MASTER.cell(row = (rows)+1, column = 1)
    c0.value = part_code
    c1 = MASTER.cell(row = (rows)+1, column = 2)
    c1.value = xx
    c2 = MASTER.cell(row = (rows)+1, column = 3)
    c2.value = vendor
    c3 = MASTER.cell(row = (rows)+1, column = 4)
    c3.value = plant
    c4 = MASTER.cell(row = (rows)+1, column = 5)
    c4.value = "QTLY"
    c5 = MASTER.cell(row = (rows)+1, column = 6)
    c5.value = '01.04.2021'
    c6 = MASTER.cell(row = (rows)+1, column = 7)
    c6.value = '31.12.9999'
    c7 = MASTER.cell(row = (rows)+1, column = 8)
    c7.value = ''
    c8 = MASTER.cell(row = (rows)+1, column = 9)
    c8.value = CONV_COST


    wb.save(r"D:\SequelString\[pcr][HCL]\AUTOFIT_MASTER.xlsx")

def hierarchy(Hierarchy,wb2,path3,part_code, vendor, plant):
    rows = Hierarchy.max_row
    
    c1a = Hierarchy.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = Hierarchy.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = Hierarchy.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = Hierarchy.cell(row = (rows)+1, column = 5)
    c4a.value = '01.04.2021'
    c5a = Hierarchy.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = Hierarchy.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = Hierarchy.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'

    wb2.save(r"D:\SequelString\[pcr][HCL]\AUTOFIT_Hierarchy.xlsx")
             

def oe(Hierarchy,wb2,path3,part_code, vendor, plant,subpart):
    hierarchy(Hierarchy,wb2,path3,part_code, vendor, plant)
    rows = Hierarchy.max_row
    c8a = Hierarchy.cell(row = (rows), column = 9)
    c8a.value = subpart
    c9a = Hierarchy.cell(row = (rows), column = 10)
    c9a.value = 'VTV'

    

    wb2.save(r"D:\SequelString\[pcr][HCL]\AUTOFIT_Hierarchy.xlsx")    




for i in range (5,13):
    
    slno = sheet_1['B'+str(i)].value
    part_code = sheet_1['G'+str(i)].value
    vendor = sheet_1['F'+str(i)].value
    plant = sheet_1['E'+str(i)].value
    partno=sheet_1['C'+str(i)].value
    print(slno)


    print('Vendor:',vendor)
    print('Plant:',plant)
    print('partcode:',part_code)
    print('partno:',partno)
    hierarchy(Hierarchy,wb2,path3,part_code, vendor, plant)

    

    for j in range (5,14):       
        concat = str(partno)+str(part_code)
        if (part_code == sheet5_1['E'+str(j)].value) and (concat == sheet5_1['G'+str(j)].value):
            CONV_COST = sheet5_1['I'+str(j)].value
            print('conv_cost',CONV_COST)
            master(MASTER,wb,path2,part_code, vendor,plant,'CONV_COST',CONV_COST)
            

    for x in range(5,138):
        if (part_code == sheet3_1['D'+str(x)].value):
            subpart = sheet3_1['E'+str(x)].value
            NO_OFF = sheet3_1['G'+str(x)].value
            
            DEPT_COST = None

            for b in range(4,61):
                if subpart == sheet4_1['C'+str(b)].value:
                    DEPT_COST = sheet4_1['U'+str(b)].value
                    print('dept_cost',DEPT_COST)
                    

            if part_code != subpart:               
                FINALPART = part_code +'_'+ subpart
                oe(Hierarchy,wb2,path3,part_code, vendor, plant,subpart)
            else:
                FINALPART = part_code
                dept_cost = None
                
            print('subpart:',subpart)
            print('no_off:',NO_OFF)
            print('finalpart:',FINALPART)
            if DEPT_COST != None:               
                master(MASTER,wb,path2,FINALPART, vendor,plant,'DEPT_COST',DEPT_COST)
            master(MASTER,wb,path2,FINALPART, vendor,plant,'NO_OFF',NO_OFF)
                            
                                


            
            
