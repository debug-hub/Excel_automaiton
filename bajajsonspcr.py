import openpyxl
from openpyxl import *
import pdf_parser
from pdf_parser import data_extractor_alphanumeric
import re


path = r"D:\SequelString\[pcr][HCL]\BAJAJSONS HARIDWAR_01.04.2021.xlsx"
excel = load_workbook(path)
excel_1 = load_workbook(path,data_only=True)

############################

path2 = r"D:\SequelString\[pcr][HCL]\BAJAJMASTER.xlsx"
wb = Workbook()
sheet2_1= wb['Sheet']
sheet2_1.title= "Master BOM Hier"
sheet2_1['A1']="BOM HIERARCHY"
sheet2_1['B1']="MASTER"
sheet2_1['C1']="DIRECT SUP"
sheet2_1['D1']="PLANT"
sheet2_1['E1']="FREQUENCY" 
sheet2_1['F1']="FROM DATE"
sheet2_1['G1']="TO DATE"
sheet2_1['H1']="PURCHASE Group"
sheet2_1['I1']="VALUE"
sheet2_1['J1']="PERCENATGE"


############################

path3 = "D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx"
wb2 = Workbook()
sheet3_1 = wb2['Sheet']

sheet3_1['A1']="BOM HIERARCHY"
sheet3_1['B1']="DIRECT SUP"
sheet3_1['C1']="PURCHASE Group"
sheet3_1['D1']="PLANT"
sheet3_1['E1']="FROM DATE"
sheet3_1['F1']="TO DATE"
sheet3_1['G1']="partcode"
sheet3_1['H1']="partcode type1(OE)"
sheet3_1['I1']="partcode Level2"
sheet3_1['J1']="partcode type2(RM/BOP/VTV/INM)"
sheet3_1['K1']="partcode Level3"
sheet3_1['L1']="partcode type3(RM/BOP/VTV/INM)"



############################

sheet = excel['TOP SHEET 01-04-21_100561']
sheet_1 = excel_1['TOP SHEET 01-04-21_100561']

sheet2 = excel['Bajajsons UTR 01.04.2021REV1']
sheet2_2 = excel_1['Bajajsons UTR 01.04.2021REV1']

############################



row_count = sheet.max_row
col_count = sheet.max_column



def master(ws_m,wb,path2,part_code, vendor, plant,date,INM,RM,gross_weight,x):
    rows = ws_m.max_row
    if INM == None:
        INM = ''
    if RM == None:
        RM = ''    
    if INM == '' and RM == '':
        finalpartcode = str(part_code)   
    elif INM != '' and RM == '':
        finalpartcode = str(part_code)+"_"+str(INM)
    elif INM == '' and RM != '':
        finalpartcode = str(part_code)+"_"+str(RM)
    else:
        finalpartcode = str(part_code)+"_"+str(INM)+"_"+str(RM)
        
    c0 = ws_m.cell(row = (rows)+1, column = 1)
    c0.value = finalpartcode
    c1 = ws_m.cell(row = (rows)+1, column = 2)
    c1.value = x
    c2 = ws_m.cell(row = (rows)+1, column = 3)
    c2.value = vendor
    c3 = ws_m.cell(row = (rows)+1, column = 4)
    c3.value = plant
    c4 = ws_m.cell(row = (rows)+1, column = 5)
    c4.value = "QTLY"
    c5 = ws_m.cell(row = (rows)+1, column = 6)
    c5.value = date
    c6 = ws_m.cell(row = (rows)+1, column = 7)
    c6.value = '31.12.9999'
    c7 = ws_m.cell(row = (rows)+1, column = 8)
    c7.value = ''
    c8 = ws_m.cell(row = (rows)+1, column = 9)
    c8.value = float(gross_weight)


    wb.save(r"D:\SequelString\[pcr][HCL]\BAJAJMASTER.xlsx")

def oe(part_code,vendor,plant,date):
    rows = sheet3_1.max_row
    print("rows"+str(rows))
    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'

    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")
      

def hierarchy(part_code, vendor, plant, date, INM,RM):
    rows = sheet3_1.max_row
    
    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+1, column = 9)
    c8a.value = INM
    c9a = sheet3_1.cell(row = (rows)+1, column = 10)
    c9a.value = "INM"
   
    c1a = sheet3_1.cell(row = (rows)+2, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+2, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+2, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+2, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+2, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+2, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+2, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+2, column = 9)
    c8a.value = INM
    c9a = sheet3_1.cell(row = (rows)+2, column = 10)
    c9a.value = "INM"
    c10a = sheet3_1.cell(row = (rows)+2, column = 11)
    c10a.value = RM
    c11a = sheet3_1.cell(row = (rows)+2, column = 12)
    c11a.value = "RM"
    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")
    print("INM saved")


def RM_H(part_code, vendor, plant, date,RM):
    rows = sheet3_1.max_row

    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+1, column = 9)
    c8a.value = RM
    c9a = sheet3_1.cell(row = (rows)+1, column = 10)
    c9a.value = "RM"
    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")


def RM_SCRAP(part_code, vendor, plant, date,RM):
    rows = sheet3_1.max_row

    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+1, column = 9)
    c8a.value = RM
    c9a = sheet3_1.cell(row = (rows)+1, column = 10)
    c9a.value = "RM"
    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")

    

def INM_SCRAP(part_code, vendor, plant, date, INM,RM):
    rows = sheet3_1.max_row

    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+1, column = 9)
    c8a.value = INM
    c9a = sheet3_1.cell(row = (rows)+1, column = 10)
    c9a.value = "INM"
    c10a = sheet3_1.cell(row = (rows)+1, column = 11)
    c10a.value = RM
    c11a = sheet3_1.cell(row = (rows)+1, column = 12)
    c11a.value = "RM"

    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")

def BOP_H(part_code, vendor, plant, date,BOP):
    rows = sheet3_1.max_row

    c1a = sheet3_1.cell(row = (rows)+1, column = 2)
    c1a.value = vendor
    c2a = sheet3_1.cell(row = (rows)+1, column = 3)
    c2a.value = ''
    c3a = sheet3_1.cell(row = (rows)+1, column = 4)
    c3a.value = plant
    c4a = sheet3_1.cell(row = (rows)+1, column = 5)
    c4a.value = date
    c5a = sheet3_1.cell(row = (rows)+1, column = 6)
    c5a.value = '31.12.9999'
    c6a = sheet3_1.cell(row = (rows)+1, column = 7)
    c6a.value = part_code
    c7a = sheet3_1.cell(row = (rows)+1, column = 8)
    c7a.value = 'OE'
    c8a = sheet3_1.cell(row = (rows)+1, column = 9)
    c8a.value = BOP
    c9a = sheet3_1.cell(row = (rows)+1, column = 10)
    c9a.value = "BOP"

    wb2.save(r"D:\SequelString\[pcr][HCL]\BAJAJHierarchy.xlsx")    
    


for i in range (3,20):
    
    slno = sheet['B'+str(i)].value
    part_code = sheet['G'+str(i)].value
    vendor = sheet['F'+str(i)].value
    plant = sheet['E'+str(i)].value
    date = '01.04.2021'
    rpl = sheet['M'+str(i)].value
    rpv = sheet_1['M'+str(i)].value

    print(slno)
    oe(part_code,vendor,plant,date)

    print('Vendor:',vendor)
    print('Plant:',plant)
    print('partcode:',part_code)
    print('DATE:',date)
    print('revised price lookup:',rpl)
    print('revised price value:',rpv)
    if '-' not in part_code:
                part_code=part_code[:5]+'-'+part_code[5:8]+'-'+part_code[8:12]+'0'
    

    for j in range (48,148):  
        partdisc=sheet2['B'+str(j)].value
        totalprice=sheet2['AI'+str(j)].value
        lvl2=sheet2_2['H'+str(j)].value



        if part_code == partdisc:
                while sheet2_2['H'+str(j)].value != None: 
                    sno = sheet2['A'+str(j)].value                    
                    pv = sheet2_2['AG'+str(j)].value 
                    subpart = sheet2_2['G'+str(j)].value
                    RM = sheet2_2['H'+str(j)].value
                    INM = sheet2_2['G'+str(j)].value
                    scraprate = sheet2['Z'+str(j)].value
                    
                    print('RM :',RM)

                    if scraprate != None:
                        scrap = sheet2_2['B'+re.sub("\D+","",scraprate)].value
                        print('scrap',scrap)
                        
                     

                    

##                    if sheet2_2['K'+str(j)].value != None and sheet2_2['I'+str(j)].value !=None:
##                        print('INM :',INM)
##                        GROSS_WT = sheet2_2['I'+str(j)].value
##                        NET_WT = sheet2_2['K'+str(j)].value
##                        print('GROSS_WT ',GROSS_WT)
##                        print('NET_WT ',NET_WT)
##
##                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,RM,GROSS_WT, 'GROSS_WT')
##                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,RM,NET_WT, 'NET_WT')
##
##                        PLATING1= sheet2['AG'+str(j)].value
##                        if PLATING1 != None:
##                            PP=PLATING1.split('*')
##                            for p in PP:
##                                
##                                if 'H' in p:
##                                    if int(re.sub('\D+','',p))<45:
##                                        PLATING_PRICE =  sheet2_2[p].value
##                                        print('PLATING_PRICE',PLATING_PRICE)
##                                        break                                      
##
##                                    
##                                else:
##                                    if 'K' not in p:
##                                        PLATING_PRICE = p
##                                        print('PLATING_PRICE',PLATING_PRICE)
##                                        print ('PP:::::::::::::',p)                           
##                            master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,RM,PLATING_PRICE, 'PLATING_PRICE')
##
##
##
##                        
##                        if INM == None:
##                            RM_H(part_code, vendor, plant, date,RM)
##                            RM_SCRAP(part_code, vendor, plant, date,scrap)
##
##                        else:
##                            hierarchy(part_code, vendor, plant, date, INM,RM)                       
##                            INM_SCRAP(part_code, vendor, plant, date, INM,scrap)



                 


                    if sheet2_2['U'+str(j)].value != None:
                        Conversion_Cost = sheet2_2['U'+str(j)].value
                        print('CONVERSION_COST',Conversion_Cost)
                        master(sheet2_1,wb,path2,part_code, vendor, plant,date,'','',Conversion_Cost,'CONVERSION_COST')


                    elif sheet2_2['G'+str(j)].value == 'PLATING':
                        PLATING1= sheet2['AG'+str(j)].value
                        if PLATING1 != None:
                            PP=PLATING1.split('*')
                            for p in PP:
                                
                                if 'H' in p:
                                    if int(re.sub('\D+','',p))<45:
                                        PLATING_PRICE =  sheet2_2[p].value
                                        print('PLATING_PRICE',PLATING_PRICE)
                                        break                                      
                                    
                                    
                                else:
                                    if 'K' not in p:
                                        PLATING_PRICE = p
                                        print('PLATING_PRICE',PLATING_PRICE)
                                                 
                        print('Subpart:',subpart)
                        print('PLATING_PRICE :',PLATING_PRICE)
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,RM,PLATING_PRICE, 'PLATING_PRICE')
                                

                    elif sheet2_2['G'+str(j)].value == 'Welding':
                        WELDING_INCH = sheet2_2['H'+str(j)].value
                        WELDING_PRICE = sheet2_2['Y'+str(j)].value
                        print('Subpart',subpart)
                        print('WELDING_INCH ',WELDING_INCH)
                        print('WELDING_PRICE ',WELDING_PRICE)
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',WELDING_INCH, 'WELDING_INCH')
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',WELDING_PRICE, 'WELDING_PRICE')

                    elif sheet2_2['G'+str(j)].value == 'Tacking':
                        TACKING = sheet2_2['H'+str(j)].value
                        TACKING_RATES = sheet2_2['Y'+str(j)].value
                        TN = sheet2['Y'+str(j)].value
                        TACKING_NUMBER = re.sub("\=\d+\.\d+\/","",TN)
                        print('Subpart',subpart)
                        print('TACKING',TACKING)
                        print('TACKING_RATES',TACKING_RATES)
                        print('TACKING_NUMBER',TACKING_NUMBER)
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',TACKING,'TACKING')
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',TACKING_RATES,'TACKING_RATES')
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',TACKING_NUMBER,'TACKING_NUMBER')

                    elif sheet2_2['G'+str(j)].value == 'Powder Coating':
                        POWDER_COAT_SURF_AREA = sheet2_2['H'+str(j)].value
                        PV = sheet2['T'+str(j)].value
                        POWDER_COAT_RATE = re.sub("\D+\d+\*","",PV)
                        print('Subpart:',subpart)
                        print('POWDER_COAT_SURF_AREA :',POWDER_COAT_SURF_AREA )
                        print('POWDER_COAT_RATE :',POWDER_COAT_RATE)
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',POWDER_COAT_SURF_AREA,'POWDER_COAT_SURF_AREA')
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',POWDER_COAT_RATE,'POWDER_COAT_RATE')

                    elif sheet2_2['G'+str(j)].value == 'SPOT WELDING':
                        WELDING_SPOTS = sheet2_2['H'+str(j)].value
                        SPOT_WELDING_RATE =sheet2_2['AE'+str(j)].value                        
                        print('Subpart:',subpart)
                        print('WELDING_SPOTS :',WELDING_SPOTS)
                        print('SPOT_WELDING_RATE :',SPOT_WELDING_RATE)
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',WELDING_SPOTS,'WELDING_SPOTS')
                        master(sheet2_1,wb,path2,part_code, vendor, plant, date,'','',SPOT_WELDING_RATE,'SPOT_WELDING_RATE')
                        
                        

        



                    else:
                        if sheet2_2['Z'+str(j)].value == None:
                            BOP = sheet2_2['G'+str(j)].value
                            BOP_value = sheet2['AA'+str(j)].value
                            no_off = sheet2_2['H'+str(j)].value
                            if 'NOS' in str(no_off):                            
                                no_off = re.sub("\D+","",no_off)    
                            print('NO_OFF:',no_off)
  

                        
                                    
                            if 'Nut Spl 6 mm' == BOP:
                                if '*' in str(BOP_value):
                                    BPV=BOP_value.split('*')
                                    for b in BPV:
                                        if 'Q' in b:
                                            BOP_COST = sheet2_2['Q'+re.sub("\D+","",b)].value   
                                else:
                                    BOP_COST = sheet2_2['Q'+re.sub("\D+","",BOP_value)].value
                                    
                                
                            elif 'NUT HEX 6 MM' == BOP:
                                if '*' in str(BOP_value):
                                    BPV=BOP_value.split('*')
                                    for b in BPV:
                                        if 'Q' in b:
                                            BOP_COST = sheet2_2['Q'+re.sub("\D+","",b)].value 
                                               
                                else:
                                    BOP_COST = sheet2_2['Q'+re.sub("\D+","",BOP_value)].value
                                    



                                
                            print('BOP_COST:',BOP_COST)
                            print('BOP:',BOP)
                            BOP_H(part_code, vendor, plant, date,BOP)
                            master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,'',no_off,'NO_OFF')
                            master(sheet2_1,wb,path2,part_code, vendor, plant, date,INM,'',BOP_COST,'BOP_COST')
                        
                        

                    j=j+1 

                        
 
            
                   
                    
                    
            

    

    

